"""Shared live-app drivers for Stage 4 numerical characterisation.

Each `drive_*` function creates its own fresh `AppTest`, drives the live
app through one analytical scenario via real widget interactions (no
internal function is called directly — matching/validation strategies are
inlined in geotestmatch.py's UI code, so AppTest is the only way to
exercise them faithfully), and returns a JSON-safe dict of the captured
results.

Used by both `scripts/update_numerical_goldens.py` (writes goldens) and
`tests/test_numerical_characterisation.py` (compares live output against
goldens), so the drive sequence lives in exactly one place.
"""

from __future__ import annotations

import datetime
import math
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd
from streamlit.testing.v1 import AppTest

REPO_ROOT = Path(__file__).resolve().parents[2]
APP_PATH = str(REPO_ROOT / "geotestmatch.py")
RUN_TIMEOUT = 180


def _to_jsonable(value: Any) -> Any:
    """Recursively convert numpy/pandas scalars and containers to plain,
    JSON-serialisable Python values. NaN passes through as float('nan')
    (Python's json module emits it as the non-standard `NaN` token, same
    as elsewhere in this repo's golden files)."""
    if isinstance(value, dict):
        return {k: _to_jsonable(v) for k, v in value.items()}
    if isinstance(value, (list, tuple)):
        return [_to_jsonable(v) for v in value]
    if isinstance(value, (np.floating,)):
        return float(value)
    if isinstance(value, (np.integer,)):
        return int(value)
    if isinstance(value, (np.bool_,)):
        return bool(value)
    if isinstance(value, (pd.Timestamp, datetime.date, datetime.datetime)):
        return str(value)
    if isinstance(value, float) and math.isnan(value):
        return value
    if isinstance(value, np.ndarray):
        return [_to_jsonable(v) for v in value.tolist()]
    return value


def _app_messages(app: AppTest) -> dict:
    return {
        "warnings": [w.value for w in app.warning],
        "errors": [e.value for e in app.error],
        "exception": app.exception is not None and len(app.exception) > 0,
    }


def _sget(ss, key, default=None):
    return ss[key] if key in ss else default


def _new_app() -> AppTest:
    app = AppTest.from_file(APP_PATH)
    app.run(timeout=RUN_TIMEOUT)
    return app


def _pick_test_auto_match(app: AppTest, test_region_prefix: str) -> None:
    """Drive Tab 1 into 'Pick Test, Auto-Match Controls' mode with one test
    region selected (by label prefix, e.g. "Aberdeen City")."""
    setup_radio = [r for r in app.radio if r.label == "Setup Mode"][0]
    setup_radio.set_value(setup_radio.options[1])
    app.run(timeout=RUN_TIMEOUT)

    ms = [m for m in app.multiselect if m.label == "select_geographies"][0]
    label = next(o for o in ms.options if o.startswith(test_region_prefix + " ("))
    ms.set_value([label])
    app.run(timeout=RUN_TIMEOUT)


def _run_match(app: AppTest) -> None:
    btn = [b for b in app.button if "Run Match Analysis" in b.label][0]
    btn.click()
    app.run(timeout=RUN_TIMEOUT)


def _structural_summary(app: AppTest) -> dict:
    ss = app.session_state
    mrm = _sget(ss, "match_run_metrics") or {}
    fc = _sget(ss, "final_controls")
    geo_col = (
        "Local Authority Area"
        if "Local Authority Area" in (fc.columns if fc is not None else [])
        else None
    )
    return _to_jsonable(
        {
            "test_regions": _sget(ss, "selected_experiment_regions"),
            "controls": sorted(fc[geo_col].tolist()) if fc is not None and geo_col else None,
            "control_count": mrm.get("control_group_size"),
            "weighted_structural_distance": mrm.get("weighted_structural_distance"),
            "mean_abs_smd": mrm.get("mean_abs_smd"),
            "smd_list": mrm.get("smd_list"),
            "best_n": _sget(ss, "best_n"),
            "match_mode_res": _sget(ss, "match_mode_res"),
            **_app_messages(app),
        }
    )


# ---------------------------------------------------------------------------
# 1-2. Structural Basic / Intermediate
# ---------------------------------------------------------------------------


def drive_structural_basic(test_region: str = "Aberdeen City") -> dict:
    app = _new_app()
    _pick_test_auto_match(app, test_region)
    _run_match(app)
    return _structural_summary(app)


def drive_structural_intermediate(test_region: str = "Aberdeen City") -> dict:
    app = _new_app()
    strategy_radio = [r for r in app.sidebar.radio if r.label == "Strategy"][0]
    strategy_radio.set_value("Intermediate (Balanced)")
    app.run(timeout=RUN_TIMEOUT)
    _pick_test_auto_match(app, test_region)
    _run_match(app)
    return _structural_summary(app)


# ---------------------------------------------------------------------------
# 3. Constraints
# ---------------------------------------------------------------------------


def _geo_of(label: str) -> str:
    return label.rsplit(" (", 1)[0]


def drive_constraints() -> dict:
    """Exercise force-include/force-exclude for both test and control groups
    in one 'Set Rules & Auto-Build Groups' run, plus a same-run attempt at
    the documented include/exclude conflict."""
    app = _new_app()
    setup_radio = [r for r in app.radio if r.label == "Setup Mode"][0]
    setup_radio.set_value(setup_radio.options[2])
    app.run(timeout=RUN_TIMEOUT)

    exp_include = [m for m in app.multiselect if m.label == "exp_include"][0]
    forced_test_label = exp_include.options[0]
    exp_include.set_value([forced_test_label])
    app.run(timeout=RUN_TIMEOUT)

    exp_exclude = [m for m in app.multiselect if m.label == "exp_exclude"][0]
    excluded_test_label = exp_exclude.options[0]
    exp_exclude.set_value([excluded_test_label])
    app.run(timeout=RUN_TIMEOUT)

    ctrl_include = [m for m in app.multiselect if m.label == "ctrl_include"][0]
    forced_ctrl_label = ctrl_include.options[0]
    ctrl_include.set_value([forced_ctrl_label])
    app.run(timeout=RUN_TIMEOUT)

    ctrl_exclude = [m for m in app.multiselect if m.label == "ctrl_exclude"][0]
    excluded_ctrl_label = ctrl_exclude.options[0]
    ctrl_exclude.set_value([excluded_ctrl_label])
    app.run(timeout=RUN_TIMEOUT)

    forced_test = _geo_of(forced_test_label)
    excluded_test = _geo_of(excluded_test_label)
    forced_ctrl = _geo_of(forced_ctrl_label)
    excluded_ctrl = _geo_of(excluded_ctrl_label)

    _run_match(app)
    ss = app.session_state
    test_regions = _sget(ss, "selected_experiment_regions") or []
    fc = _sget(ss, "final_controls")
    control_regions = fc["Local Authority Area"].tolist() if fc is not None else []

    constraints_result = _to_jsonable(
        {
            "forced_test_region": forced_test,
            "forced_test_region_in_test_group": forced_test in test_regions,
            "excluded_test_region": excluded_test,
            "excluded_test_region_in_test_group": excluded_test in test_regions,
            "forced_control_region": forced_ctrl,
            "forced_control_region_in_control_group": forced_ctrl in control_regions,
            "excluded_control_region": excluded_ctrl,
            "excluded_control_region_in_control_group": excluded_ctrl in control_regions,
            "n_test_regions": len(test_regions),
            "n_control_regions": len(control_regions),
            "guided_share_info": _sget(ss, "guided_share_info"),
            **_app_messages(app),
        }
    )

    # ---- Conflict sub-scenario: attempt to force-include the same region into
    # both the test group and the control group within one run. ----
    conflict_app = _new_app()
    setup_radio = [r for r in conflict_app.radio if r.label == "Setup Mode"][0]
    setup_radio.set_value(setup_radio.options[2])
    conflict_app.run(timeout=RUN_TIMEOUT)

    ctrl_include = [m for m in conflict_app.multiselect if m.label == "ctrl_include"][0]
    target_label = ctrl_include.options[0]
    ctrl_include.set_value([target_label])
    conflict_app.run(timeout=RUN_TIMEOUT)

    exp_include = [m for m in conflict_app.multiselect if m.label == "exp_include"][0]
    exp_include.set_value([target_label])
    conflict_app.run(timeout=RUN_TIMEOUT)

    _run_match(conflict_app)

    ctrl_include_after = [m for m in conflict_app.multiselect if m.label == "ctrl_include"][0]
    exp_include_after = [m for m in conflict_app.multiselect if m.label == "exp_include"][0]

    conflict_result = _to_jsonable(
        {
            "target_region": _geo_of(target_label),
            # Streamlit silently drops a persisted multiselect value once it's no
            # longer in that widget's (freshly recomputed) options list, rather
            # than raising — so ctrl_include's stale selection is cleared once
            # exp_include claims the same region, and the app's own conflict
            # error path (st.error + st.stop on overlapping constraints) is
            # never reached via this sequence of live widget interactions.
            "ctrl_include_value_after_conflict_attempt": list(ctrl_include_after.value),
            "exp_include_value_after_conflict_attempt": list(exp_include_after.value),
            **_app_messages(conflict_app),
        }
    )

    return {"constraints": constraints_result, "conflict": conflict_result}


# ---------------------------------------------------------------------------
# 3b. Global exclusion (exclude-from-both control)
# ---------------------------------------------------------------------------


def drive_global_exclusion() -> dict:
    """Exercise the 'exclude from both' global-exclusion multiselect in one
    'Set Rules & Auto-Build Groups' run: selects one region, drives two more
    reruns (an unrelated rerun, then the Run button click) to prove the
    selection survives, then checks the region is absent from the frozen
    control candidate pool (not just the final control group) and from the
    final test group."""
    app = _new_app()
    setup_radio = [r for r in app.radio if r.label == "Setup Mode"][0]
    setup_radio.set_value(setup_radio.options[2])
    app.run(timeout=RUN_TIMEOUT)

    global_exclude = [m for m in app.multiselect if m.label == "global_exclude"][0]
    excluded_label = global_exclude.options[0]
    excluded_geo = _geo_of(excluded_label)
    global_exclude.set_value([excluded_label])
    app.run(timeout=RUN_TIMEOUT)

    # An unrelated rerun (re-running with no new widget interaction) — the
    # selection must still be there afterwards, not just immediately after
    # the multiselect's own on_change rerun.
    app.run(timeout=RUN_TIMEOUT)
    global_exclude_mid = [m for m in app.multiselect if m.label == "global_exclude"][0]
    value_before_run_click = list(global_exclude_mid.value)

    _run_match(app)

    global_exclude_after = [m for m in app.multiselect if m.label == "global_exclude"][0]
    value_after_run_click = list(global_exclude_after.value)

    ss = app.session_state
    snapshot = _sget(ss, "match_run_snapshot") or {}
    test_regions = _sget(ss, "selected_experiment_regions") or []
    fc = _sget(ss, "final_controls")
    control_regions = fc["Local Authority Area"].tolist() if fc is not None else []

    return _to_jsonable(
        {
            "excluded_geo": excluded_geo,
            "value_persisted_before_run_click": value_before_run_click == [excluded_label],
            "value_persisted_after_run_click": value_after_run_click == [excluded_label],
            "excluded_from_control_candidate_pool": excluded_geo
            not in snapshot.get("control_pool_geos", []),
            "excluded_from_final_test_group": excluded_geo not in test_regions,
            "excluded_from_final_control_group": excluded_geo not in control_regions,
            "snapshot_global_exclusions": snapshot.get("global_exclusions", []),
            **_app_messages(app),
        }
    )


# ---------------------------------------------------------------------------
# 4. KPI pattern
# ---------------------------------------------------------------------------


def drive_kpi_pattern(tmp_path: Path) -> dict:
    from tests.fixture_factories.write_aggregated_kpi_xlsx import write_aggregated_kpi_xlsx

    regions = ["RegionA", "RegionB", "RegionC", "RegionD", "RegionE"]
    kpi_path = write_aggregated_kpi_xlsx(
        tmp_path / "kpi_pattern.xlsx",
        regions,
        aggregation_level_col="TV Region",
        n_weeks=20,
        seed=55,
    )

    app = _new_app()
    method_radio = [r for r in app.sidebar.radio if r.label == "Matching method"][0]
    method_radio.set_value("KPI Pattern")
    app.run(timeout=RUN_TIMEOUT)

    uploader = [f for f in app.sidebar.file_uploader if f.key == "kpi_pattern_sidebar_uploader"][0]
    uploader.set_value(
        (
            "kpi_pattern.xlsx",
            kpi_path.read_bytes(),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    )
    app.run(timeout=RUN_TIMEOUT)

    start_select = [s for s in app.sidebar.selectbox if s.key == "kpi_pattern_date_start_sidebar"][
        0
    ]
    end_select = [s for s in app.sidebar.selectbox if s.key == "kpi_pattern_date_end_sidebar"][0]
    period_count = len(start_select.options)
    assert period_count == len(end_select.options)

    setup_radio = [r for r in app.radio if r.label == "Setup Mode"][0]
    setup_radio.set_value(setup_radio.options[1])
    app.run(timeout=RUN_TIMEOUT)

    ms = [m for m in app.multiselect if m.label == "select_geographies"][0]
    test_label = ms.options[0]
    ms.set_value([test_label])
    app.run(timeout=RUN_TIMEOUT)

    _run_match(app)

    ss = app.session_state
    mrm = _sget(ss, "match_run_metrics") or {}
    fc = _sget(ss, "final_controls")

    return _to_jsonable(
        {
            "test_region": _sget(ss, "selected_experiment_regions"),
            "controls": sorted(fc["TV Region"].tolist()) if fc is not None else None,
            "control_count": mrm.get("control_group_size"),
            "weighted_structural_distance": mrm.get("weighted_structural_distance"),
            "mean_abs_smd": mrm.get("mean_abs_smd"),
            "period_count": period_count,
            **_app_messages(app),
        }
    )


# ---------------------------------------------------------------------------
# Shared setup for weekly / evaluation / daily scenarios: Manual Selection
# with fixed, deterministic real-workbook regions, then Tab 2/3 KPI upload.
# ---------------------------------------------------------------------------

TEST_REGION = "Aberdeen City"
CONTROL_REGIONS = ["Aberdeenshire", "Angus"]


def _manual_match(app: AppTest) -> None:
    setup_radio = [r for r in app.radio if r.label == "Setup Mode"][0]
    setup_radio.set_value(setup_radio.options[0])
    app.run(timeout=RUN_TIMEOUT)

    test_ms = [m for m in app.multiselect if m.label == "test_geos_manual"][0]
    test_label = next(o for o in test_ms.options if o.startswith(TEST_REGION + " ("))
    test_ms.set_value([test_label])
    app.run(timeout=RUN_TIMEOUT)

    ctrl_ms = [m for m in app.multiselect if m.label == "control_geos_manual"][0]
    ctrl_labels = [
        o for o in ctrl_ms.options if any(o.startswith(c + " (") for c in CONTROL_REGIONS)
    ]
    ctrl_ms.set_value(ctrl_labels)
    app.run(timeout=RUN_TIMEOUT)

    _run_match(app)


def _upload_kpi(app: AppTest, mode_prefix: str, filename: str, file_bytes: bytes) -> None:
    uploaders = [f for f in app.file_uploader if f.key.startswith(f"kpi_uploader_{mode_prefix}_")]
    uploaders[0].set_value(
        (filename, file_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    )
    app.run(timeout=RUN_TIMEOUT)


def _validation_result_summary(res: dict) -> dict:
    return _to_jsonable(
        {
            "controls": sorted(res.get("selected_regions") or []),
            "n_pre_periods": res.get("n_pre_periods"),
            "n_folds": len(res["rolling_origin_folds"])
            if res.get("rolling_origin_folds") is not None
            else None,
            "smape": res.get("smape"),
            "rmse": res.get("rmse"),
            "rolling_smape_mean": res.get("rolling_smape_mean"),
            "rolling_rmse_mean": res.get("rolling_rmse_mean"),
            "rolling_bias_pct_mean": res.get("rolling_bias_pct_mean"),
            "dw_stat": res.get("dw_stat"),
            "counterfactual_reliability": res.get("counterfactual_reliability"),
            "uplift": res.get("uplift"),
            "uplift_pct": res.get("uplift_pct"),
            "n_placebos": len(res.get("placebos") or []),
            "median_placebo_uplift": res.get("median_placebo_uplift"),
            "placebo_range_lower": res.get("placebo_range_lower"),
            "placebo_range_upper": res.get("placebo_range_upper"),
            "placebo_percentile_rank": res.get("placebo_percentile_rank"),
            "placebo_p_value_one_sided": res.get("placebo_p_value_one_sided"),
            "placebo_p_value_two_sided": res.get("placebo_p_value_two_sided"),
            "placebo_z_score": res.get("placebo_z_score"),
            "lag_periods": res.get("lag_periods"),
            "lag_drop_metadata": res.get("lag_drop_metadata"),
            "n_selected_features": res.get("n_selected_features"),
            "time_series_frequency": res.get("time_series_frequency"),
        }
    )


# ---------------------------------------------------------------------------
# 5. Weekly validation (Design mode — no test period, pre-period fit only)
# ---------------------------------------------------------------------------


def drive_weekly_validation(tmp_path: Path) -> dict:
    from tests.fixture_factories.write_correlated_kpi_xlsx import write_correlated_kpi_xlsx

    kpi_path = write_correlated_kpi_xlsx(
        tmp_path / "weekly.xlsx",
        TEST_REGION,
        CONTROL_REGIONS,
        metric_name="Sales",
        n_periods=60,
        freq="W",
        seed=123,
    )

    app = _new_app()
    _manual_match(app)
    _upload_kpi(app, "design", "weekly.xlsx", kpi_path.read_bytes())

    run_btn = [b for b in app.button if b.key == "design_run_button"][0]
    run_btn.click()
    app.run(timeout=RUN_TIMEOUT)

    ss = app.session_state
    vr = _sget(ss, "validation_results")
    res = vr["results"]["User Selected Test and Control"]
    summary = _validation_result_summary(res)
    summary.update(_app_messages(app))
    return summary


# ---------------------------------------------------------------------------
# 6. Completed-test evaluation (Evaluate mode — uplift + placebo)
# ---------------------------------------------------------------------------


def drive_completed_test_evaluation(tmp_path: Path) -> dict:
    from tests.fixture_factories.write_correlated_kpi_xlsx import write_correlated_kpi_xlsx

    kpi_path = write_correlated_kpi_xlsx(
        tmp_path / "weekly_eval.xlsx",
        TEST_REGION,
        CONTROL_REGIONS,
        metric_name="Sales",
        n_periods=60,
        freq="W",
        seed=123,
    )

    app = _new_app()
    _manual_match(app)
    _upload_kpi(app, "evaluate", "weekly_eval.xlsx", kpi_path.read_bytes())

    run_btn = [b for b in app.button if b.key == "evaluate_run_button"][0]
    run_btn.click()
    app.run(timeout=RUN_TIMEOUT)

    ss = app.session_state
    vr = _sget(ss, "validation_results")
    res = vr["results"]["User Selected Test and Control"]
    summary = _validation_result_summary(res)
    summary["actual_total"] = (
        float(np.sum(res["y_test_actual"])) if res.get("y_test_actual") is not None else None
    )
    summary["counterfactual_total"] = (
        float(np.sum(res["y_pred_test"])) if res.get("y_pred_test") is not None else None
    )
    summary.update(_app_messages(app))
    return summary


# ---------------------------------------------------------------------------
# 7. Daily evaluation
# ---------------------------------------------------------------------------


def drive_daily_evaluation(tmp_path: Path) -> dict:
    from tests.fixture_factories.write_correlated_kpi_xlsx import write_correlated_kpi_xlsx

    kpi_path = write_correlated_kpi_xlsx(
        tmp_path / "daily.xlsx",
        TEST_REGION,
        CONTROL_REGIONS,
        metric_name="Sales",
        n_periods=150,
        freq="D",
        seed=321,
    )

    app = _new_app()
    _manual_match(app)
    _upload_kpi(app, "evaluate", "daily.xlsx", kpi_path.read_bytes())

    freq_radio = [r for r in app.radio if r.key == "evaluate_time_series_frequency"][0]
    inferred_frequency_options = list(freq_radio.options)
    freq_radio.set_value("daily")
    app.run(timeout=RUN_TIMEOUT)

    lag_cb = [c for c in app.checkbox if c.key == "evaluate_include_lagged_controls"][0]
    lag_cb.set_value(True)
    app.run(timeout=RUN_TIMEOUT)

    run_btn = [b for b in app.button if b.key == "evaluate_run_button"][0]
    run_btn_disabled_before_run = run_btn.disabled
    run_btn.click()
    app.run(timeout=RUN_TIMEOUT)

    ss = app.session_state
    vr = _sget(ss, "validation_results")
    res = vr["results"]["User Selected Test and Control"]
    summary = _validation_result_summary(res)
    summary["actual_total"] = (
        float(np.sum(res["y_test_actual"])) if res.get("y_test_actual") is not None else None
    )
    summary["counterfactual_total"] = (
        float(np.sum(res["y_pred_test"])) if res.get("y_pred_test") is not None else None
    )
    summary["frequency_radio_options"] = inferred_frequency_options
    summary["run_button_disabled_before_run"] = run_btn_disabled_before_run
    summary.update(_app_messages(app))
    return summary


# ---------------------------------------------------------------------------
# 8. Time-series tracking-outage exclusion (Design mode)
# ---------------------------------------------------------------------------


def drive_outage_exclusion(tmp_path: Path) -> dict:
    """Injects one market-wide zero week into an otherwise-clean correlated
    weekly fixture, then drives Design-mode validation to prove: the outage
    week is auto-detected and preselected in the 'Periods to exclude...'
    widget, the selection survives the Run button click, and the run
    snapshot records it as both an automatic and an effective exclusion."""
    from openpyxl import load_workbook

    from tests.fixture_factories.write_correlated_kpi_xlsx import write_correlated_kpi_xlsx

    kpi_path = write_correlated_kpi_xlsx(
        tmp_path / "outage.xlsx",
        TEST_REGION,
        CONTROL_REGIONS,
        metric_name="Sales",
        n_periods=60,
        freq="W",
        seed=123,
    )

    # Zero out one interior date column for every region — a market-wide
    # tracking outage week, injected on top of otherwise-clean correlated data.
    wb = load_workbook(kpi_path)
    ws = wb.active
    header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    outage_col_idx = 30  # 0-based; well inside the 60-period pre-period history
    for row in ws.iter_rows(min_row=2):
        row[outage_col_idx].value = 0
    wb.save(kpi_path)
    outage_date = pd.Timestamp(header[outage_col_idx])

    app = _new_app()
    _manual_match(app)
    _upload_kpi(app, "design", "outage.xlsx", kpi_path.read_bytes())

    outage_widget = [m for m in app.multiselect if m.label == "kpi_outage_exclude"][0]
    preselected_labels = list(outage_widget.value)
    outage_date_preselected = any(
        lbl.startswith(outage_date.strftime("%d %b %y")) for lbl in preselected_labels
    )

    app.run(timeout=RUN_TIMEOUT)  # unrelated rerun — selection must survive
    outage_widget_mid = [m for m in app.multiselect if m.label == "kpi_outage_exclude"][0]
    value_persisted_before_run_click = list(outage_widget_mid.value) == preselected_labels

    run_btn = [b for b in app.button if b.key == "design_run_button"][0]
    run_btn.click()
    app.run(timeout=RUN_TIMEOUT)

    outage_widget_after = [m for m in app.multiselect if m.label == "kpi_outage_exclude"][0]
    value_persisted_after_run_click = list(outage_widget_after.value) == preselected_labels

    ss = app.session_state
    vr = _sget(ss, "validation_results") or {}

    summary = _to_jsonable(
        {
            "n_preselected": len(preselected_labels),
            "outage_date_preselected": outage_date_preselected,
            "value_persisted_before_run_click": value_persisted_before_run_click,
            "value_persisted_after_run_click": value_persisted_after_run_click,
            "automatic_outage_dates": vr.get("automatic_outage_dates", []),
            "manual_excluded_dates": vr.get("manual_excluded_dates", []),
            "effective_excluded_dates": vr.get("effective_excluded_dates", []),
            **_app_messages(app),
        }
    )
    return summary
